import poplib
import datetime
import email
import configparser
import docx        # 读取Word文档
import os
from email.parser import Parser
from email.header import decode_header,Header
from email.utils import parseaddr
from openpyxl import load_workbook
from openpyxl import Workbook

def is_zuoye_from_title(title):
    if title is None:
        return True
    
    if "已撤回" in title:
        return False
    
    return True

def find_student_by_email(email, wb):
    #print("查找学号:", id)
    # 遍历所有的sheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        #打印班级名称
        #print("班级名称:", ws.title)
        for row in ws.iter_rows(values_only=True):
            # 打印学号和姓名
            #将学号转换为字符串
            #
            #print("学号:", row[0], " 姓名:", row[1], " id:", id)
            if str(row[2]) == email:
                #print("找到学号:", id)
                return row[0], row[1], ws.title
            
    return None, None, None

# 此函数通过使用poplib实现接收邮件
def recv_email_by_pop3():
    #读取email.ini文件，huoquemail.ini文件中的内容
    config = configparser.ConfigParser()
    config.read('email.ini')

    # 要进行邮件接收的邮箱。改成自己的邮箱
    email_address = config.get('Email', 'address')
    email_password = config.get('Email', 'password')
    pop_server_host = config.get('Email', 'pop_server')
    # config.get('Email', 'pop_port')转为int类型
    pop_server_port = config.getint('Email', 'pop_port')

    try:
        # 连接pop服务器。如果没有使用SSL，将POP3_SSL()改成POP3()即可其他都不需要做改动
        email_server = poplib.POP3_SSL(host=pop_server_host, port=pop_server_port, timeout=10)
        print("pop3----connect server success, now will check username")
    except:
        print("pop3----sorry the given email server address connect time out")
        exit(1)
    try:
        # 验证邮箱是否存在
        email_server.user(email_address)
        print("pop3----username exist, now will check password")
    except:
        print("pop3----sorry the given email address seem do not exist")
        exit(1)
    try:
        # 验证邮箱密码是否正确
        email_server.pass_(email_password)
        print("pop3----password correct,now will list email")
    except:
        print("pop3----sorry the given username seem do not correct")
        exit(1)

    wb_student = load_workbook('学生信息.xlsx')
    wb_result = Workbook()
    ws_result = wb_result.active
    ws_result.title = "作业统计"  
    # 第一行写标题
    ws_result.append(['序号','时间', '发件人', '学号', '班级', '邮箱', '主题', '作业', '实验报告', '备注', '附件名字'])

    email_index = 0
    # 添加一个sheet

    month = 3

    #现在开始收取邮件
    email_count = len(email_server.list()[1])
    print("邮件数量:", email_count)
    # list()返回所有邮件的编号:
    resp, mails, octets = email_server.list()
    # 遍历所有的邮件
    for i in range(1, len(mails) + 1):
        # 通过retr(index)读取第index封邮件的内容；这里读取最后一封，也即最新收到的那一封邮件
        resp, lines, octets = email_server.retr(i)

        # lines是邮件内容，列表形式使用join拼成一个byte变量
        email_content = b'\r\n'.join(lines)
        try:
            # 再将邮件内容由byte转成str类型
            email_content = email_content.decode('utf-8')
        except Exception as e:
            print(str(e))
            continue

        msg = Parser().parsestr(email_content)

        ret = process_email(msg, month, ws_result, email_index, wb_student)
        if ret == 0:
            continue    
        elif ret == -1:
            break   

        email_index += 1
        #if email_index >= 10:
        #    break

    # 关闭连接
    email_server.close()
    
    # 获取当前时间并格式化成字符串
    now = datetime.datetime.now()

    wb_result.save('作业统计_' + str(month) + now.strftime('%Y%m%d%H%M%S') + '.xlsx')
    # 只读的
    wb_student.close()

def process_email(msg, month, ws_result, email_index, wb_student):
    
    #获取时间，并将时间转换为本地时间
    date1 = msg.get('Date')
    #将date1里面的内容转为标准时间变脸
    date2 = email.utils.parsedate(date1)

    # if date2[1] != month:
    #    return -1
    #print(date2)
    #比较当前时间是否大于2024年4月30日, 格式化输出当前时间
    date3 = datetime.datetime(date2[0],date2[1],date2[2],date2[3],date2[4],date2[5])
    date4 = datetime.datetime(2024,3,6,0,0,0)
    if date3 < date4:
        return 0
    emailTime = date3.strftime('%Y-%m-%d %H:%M:%S')
    emailSubJect = decode_str(msg.get('Subject'))
    if not is_zuoye_from_title(emailSubJect):
        return 0

    hdr, addr = parseaddr(msg.get('From'))
    name = decode_str(hdr)
    emailFrom = u'%s <%s>' % (name, addr)
    emailContent = ""

    stu_id, stu_name, stu_class = find_student_by_email(addr, wb_student)
    
    if (stu_id is None):
        return 0

    #将上面的信息打印出来，用一行打印出来
    print(" ==>", stu_id, email_index, emailTime, emailFrom, emailSubJect)
    # 向wb_result中写入新行
    is_work_zuoye = ''
    is_report = ''
    
    has_write_log = False
    attachment_files = []
    for part in msg.walk():
        file_name = part.get_filename()  # 获取附件名称类型
        contentType = part.get_content_type() #获取数据类型
        mycode = part.get_content_charset()  #获取编码格式
        if file_name:
            h = Header(file_name)
            dh = decode_header(h)  # 对附件名称进行解码
            filename = dh[0][0]
            if dh[0][1]:
                filename = decode_str(str(filename, dh[0][1]))  # 将附件名称可读化

            # 判断文件名必须是doc或者docx,否则循环
            if not filename.endswith('.doc') and not filename.endswith('.docx'):
                continue
            #在文件名前面加上时间，以防止重名
            timePrefix = date3.strftime('%Y%m%d%H%M%S')

            folderPrefix = stu_class + "/" + str(stu_id)
            #folderPrefix = date3.strftime('%Y%m%d')
            #判断folderPrefix目录是否存在，如果不存在就创建
            if not os.path.exists(folderPrefix):
                os.makedirs(folderPrefix)

            fullfilename = folderPrefix + "/" + timePrefix + "_" + filename
            attachment_files.append(fullfilename)
            data = part.get_payload(decode=True)  # 下载附件
            with open(fullfilename, 'wb') as f: # 在当前目录下创建文件，注意二进制文件需要用wb模式打开
            #with open('指定目录路径'+filename, 'wb') as f: 也可以指定下载目录
                f.write(data)  # 保存附件
                f.close()
            #print(f'附件 {filename} 已下载完成')
            # 看看文件名是否后缀为.docx
            filetype = docx_file_detect(fullfilename)

            is_report = ''
            is_work_zuoye = ''
            
            #根据文件名字重新命名文件，如果是实验报告就加上实验报告，如果是作业就加上作业
            if filetype == 1:
                #重新命名文件
                os.rename(fullfilename, folderPrefix + "/实验报告_" + timePrefix + "_" + filename)
                is_report = '是'
                fullfilename = folderPrefix + "/实验报告_" + timePrefix + "_" + filename
            elif filetype == 2:
                #重新命名文件
                os.rename(fullfilename, folderPrefix + "/作业_" + timePrefix + "_" + filename)
                fullfilename = folderPrefix + "/作业_" + timePrefix + "_" + filename
                is_work_zuoye = '是'

            record_row = [email_index, emailTime, stu_name, stu_id, stu_class, addr, emailSubJect, is_work_zuoye, is_report, "", fullfilename]
            ws_result.append(record_row)

            has_write_log = True

        elif contentType == 'text/plain': #or contentType == 'text/html':
            # 输出正文 也可以写入文件
            data = part.get_payload(decode=True)
            content = data.decode(mycode)
            #print('正文：',content)
            emailContent += content

    # 如果没有附件，我就认为他可能是用文本交的作业
    if not has_write_log:
        record_row = [email_index, emailTime, stu_name, stu_id, stu_class, addr, emailSubJect, '', '', "", str(attachment_files)]
        ws_result.append(record_row)

    #['时间', '发件人', '学号', '班级', '邮箱', '主题', '作业', '试验报告', '备注', '附件名字']
    
    return 1
# 分析docx文件内容，看看是不是实验报告
def docx_file_detect(filename):
    #只处理docx,其他的默认当做不认识的文件
    if not filename.endswith('.docx'):
        return 0
    
    #判断文件名里面是否包含作业或者实验报告
    if '作业' in filename:
        return 2
    if '报告' in filename:
        return 1
    
    # 打印文件名字
    # print(filename)
    doc_file =docx.Document(filename)

    paragraph_index = 0
    for paragraph in doc_file.paragraphs:
        # 搜索关键字，看看是否找到  "试验报告"
        if "实验报告" in paragraph.text:
            #print("这是一份实验报告")
            #print(paragraph.text)
            return 1
        if "作业" in paragraph.text:
            #print("这是一份作业")
            #print(paragraph.text)
            return 2
        
        paragraph_index += 1
        if paragraph_index > 10:
            break

    print("不能确认文件类型")
    # 默认是没有找到类型
    return 0

# 解码
def decode_str(s):
    value, charset = decode_header(s)[0]
    if charset:
        value = value.decode(charset)
    return value

# 猜测字符编码
def guess_charset(msg):
    # 先从msg对象获取编码:
    charset = msg.get_charset()
    if charset is None:
        # 如果获取不到，再从Content-Type字段获取:
        content_type = msg.get('Content-Type', '').lower()
        for item in content_type.split(';'):
            item = item.strip()
            if item.startswith('charset'):
                charset = item.split('=')[1]
                break
    return charset

if __name__ == "__main__":
    recv_email_by_pop3()